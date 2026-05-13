"""
Обработчик Excel-файлов.
Запись извлечённых данных контрактов в Excel в формате .xlsx.
Поддерживает создание нового файла и добавление данных в существующий.
"""

from pathlib import Path
from typing import Optional

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from parser_engine import ContractData

HEADERS = [
    "Собственник товара",
    "тип",
    "код ЛС льготный",
    "№ договора",
    "№ аукциона",
    "Международное непатентованное название (МНН)",
    "Торговое наименование",
    "Лек. Форма (форма выпуска), дозировка",
    "Ед. изм.",
    "Общее кол-во, ед. изм.",
    "Производитель",
    "Цена ед. продукции, руб. в месте поставки",
    "Общая цена, руб.",
    "поставщик",
    "способ размещения",
    "Сумма",
]

COLUMN_WIDTHS = [15, 18, 15, 30, 25, 30, 25, 50, 10, 15, 45, 20, 20, 25, 25, 20]

SHEET_NAME = "Контракты"


def _owner_abbreviation(full_name: str) -> str:
    """Сокращение наименования заказчика до аббревиатуры."""
    mapping = {
        "МИНЗДРАВ": "МЗ",
        "МИНИСТЕРСТВО ЗДРАВООХРАНЕНИЯ": "МЗ",
    }
    upper = full_name.upper()
    for key, abbr in mapping.items():
        if key in upper:
            return abbr
    return full_name


RED_FILL = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
BLUE_FILL = PatternFill(start_color="6699FF", end_color="6699FF", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")


def contract_to_row(data: ContractData) -> list:
    contract_str = data.contract_number
    if data.contract_date:
        contract_str += f" от {data.contract_date}"

    owner = _owner_abbreviation(data.customer_short_name)

    qty_value = data.quantity_all_values if data.quantity_mismatch else data.quantity_packages

    return [
        owner,
        "Основная заявка",
        "",
        contract_str,
        data.notice_number,
        data.mnn,
        data.trade_name,
        data.dosage_form,
        data.unit,
        qty_value,
        data.manufacturer,
        data.unit_price,
        data.total_price,
        data.supplier_short_name,
        data.procurement_method,
        data.total_price,
    ]


def _apply_header_style(ws, row_num: int) -> None:
    header_font = Font(bold=True, size=10)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, size=10, color="FFFFFF")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col_idx in range(1, len(HEADERS) + 1):
        cell = ws.cell(row=row_num, column=col_idx)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = center_align


def _apply_data_style(ws, row_num: int) -> None:
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    wrap_align = Alignment(vertical="center", wrap_text=True)

    for col_idx in range(1, len(HEADERS) + 1):
        cell = ws.cell(row=row_num, column=col_idx)
        cell.border = thin_border
        cell.alignment = wrap_align
        if col_idx in (10, 12, 13, 16):
            cell.number_format = "#,##0.00"


def create_new_excel(file_path: str) -> str:
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME

    for col_idx, header in enumerate(HEADERS, 1):
        ws.cell(row=1, column=col_idx, value=header)

    _apply_header_style(ws, 1)

    for col_idx, width in enumerate(COLUMN_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    path = file_path
    if not path.endswith(".xlsx"):
        path += ".xlsx"

    wb.save(path)
    return path


def write_contracts_to_excel(
    file_path: str,
    contracts: list[ContractData],
    sheet_name: Optional[str] = None,
) -> tuple[bool, str]:
    path = Path(file_path)
    target_sheet = sheet_name or SHEET_NAME

    try:
        if path.exists() and path.suffix == ".xlsx":
            wb = load_workbook(str(path))
            if target_sheet in wb.sheetnames:
                ws = wb[target_sheet]
            else:
                ws = wb.create_sheet(target_sheet)
                for col_idx, header in enumerate(HEADERS, 1):
                    ws.cell(row=1, column=col_idx, value=header)
                _apply_header_style(ws, 1)
                for col_idx, width in enumerate(COLUMN_WIDTHS, 1):
                    ws.column_dimensions[get_column_letter(col_idx)].width = width
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = target_sheet
            for col_idx, header in enumerate(HEADERS, 1):
                ws.cell(row=1, column=col_idx, value=header)
            _apply_header_style(ws, 1)
            for col_idx, width in enumerate(COLUMN_WIDTHS, 1):
                ws.column_dimensions[get_column_letter(col_idx)].width = width

        # Find last row with data
        last_row = ws.max_row
        if last_row == 1:
            first_cell = ws.cell(row=1, column=1).value
            if first_cell is None:
                last_row = 0

        # Write contracts
        written = 0
        qty_mismatch_rows = []
        dosage_mnn_only_rows = []
        for contract in contracts:
            row_data = contract_to_row(contract)
            row_num = last_row + 1 + written
            for col_idx, value in enumerate(row_data, 1):
                ws.cell(row=row_num, column=col_idx, value=value)
            _apply_data_style(ws, row_num)
            if contract.quantity_mismatch:
                qty_mismatch_rows.append(row_num)
            if contract.dosage_form_mnn_only:
                dosage_mnn_only_rows.append(row_num)
            written += 1

        # Check for duplicate rows in the entire sheet
        all_rows = {}
        duplicate_rows = set()
        for row_num in range(2, ws.max_row + 1):
            row_values = tuple(
                ws.cell(row=row_num, column=col).value
                for col in range(1, len(HEADERS) + 1)
            )
            if row_values in all_rows:
                duplicate_rows.add(row_num)
                duplicate_rows.add(all_rows[row_values])
            else:
                all_rows[row_values] = row_num

        has_duplicates = len(duplicate_rows) > 0
        if has_duplicates:
            for row_num in duplicate_rows:
                for col_idx in range(1, len(HEADERS) + 1):
                    ws.cell(row=row_num, column=col_idx).fill = BLUE_FILL

        # Apply red fill for quantity mismatches (on top of blue if needed)
        for row_num in qty_mismatch_rows:
            ws.cell(row=row_num, column=10).fill = RED_FILL

        # Apply yellow fill for dosage_form with MNN only
        for row_num in dosage_mnn_only_rows:
            ws.cell(row=row_num, column=8).fill = YELLOW_FILL

        save_path = str(path)
        if not save_path.endswith(".xlsx"):
            save_path = save_path.rsplit(".", 1)[0] + ".xlsx"

        wb.save(save_path)
        msg = f"Записано {written} контракт(ов) в {save_path}"
        if has_duplicates:
            msg += "\n\nПри выгрузке обнаружены дубли (выделены синим цветом)"
        return True, msg

    except PermissionError:
        return False, (
            "Файл занят другой программой. "
            "Закройте Excel и повторите попытку."
        )
    except Exception as e:
        return False, f"Ошибка записи: {e}"


def get_sheet_names(file_path: str) -> list[str]:
    try:
        path = Path(file_path)
        if path.suffix == ".xlsx":
            wb = load_workbook(str(path), read_only=True)
            return wb.sheetnames
        elif path.suffix == ".xls":
            import xlrd
            wb = xlrd.open_workbook(str(path))
            return wb.sheet_names()
    except Exception:
        pass
    return []
